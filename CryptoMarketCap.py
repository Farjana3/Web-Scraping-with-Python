from logging import exception
from bs4 import BeautifulSoup
import openpyxl
import requests

excel = openpyxl.Workbook()

print(excel.sheetnames)
sheet = excel.active
sheet.title='CryptoSheet'
print(excel.sheetnames)
sheet.append(['Rank','Name','Price','1h','24h','7d','Volume24h','Mkt Cap'])

try:
    source = requests.get('https://www.coingecko.com/')
    source.raise_for_status()

    soup = BeautifulSoup(source.text,'html.parser')
    coins = soup.find('tbody').find_all('tr')

    for coin in coins:
        
        rank= coin.find('td',class_="table-number tw-text-left text-xs cg-sticky-col cg-sticky-second-col tw-max-w-14 lg:tw-w-14").get_text(strip=True)
        name = coin.find('td',class_="py-0 coin-name cg-sticky-col cg-sticky-third-col px-0").a.span.get_text(strip=True)
        price = coin.find('td',class_="td-price price text-right pl-0").span.text
        hour1 = coin.find('td',class_="td-change1h change1h stat-percent text-right col-market").span.text
        hour24 = coin.find('td',class_="td-change24h change24h stat-percent text-right col-market").span.text
        day7= coin.find('td',class_="td-change7d change7d stat-percent text-right col-market").span.text
        volume24h= coin.find('td',class_="td-liquidity_score lit text-right col-market").span.text
        mktcap=coin.find('td',class_="td-market_cap cap col-market cap-price text-right").span.text
        print(rank,name,price,hour1,hour24,day7,volume24h,mktcap)
        sheet.append([rank,name,price,hour1,hour24,day7,volume24h,mktcap] )     


except Exception as e:
    print(e)

excel.save('Crypto Market Cap.xlsx')

